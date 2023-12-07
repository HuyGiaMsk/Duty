import datetime
import os
import time
from datetime import datetime
from enum import Enum
from logging import Logger
from typing import Dict, Tuple

from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement

from src.AutomatedTask import AutomatedTask
from src.FileUtil import get_excel_data_in_column_start_at_row
from src.ThreadLocalLogger import get_current_logger


class BookingToInfoIndex(Enum):
    FCR_INDEX_IN_TUPLE = 0
    STATUS_INDEX_IN_TUPLE = 1


class Duty(AutomatedTask):
    fcr_to_file_remane = {}

    def __init__(self, settings: dict[str, str]):
        super().__init__(settings)
        self._document_folder = self._download_folder

    def mandatory_settings(self) -> set[str]:
        mandatory_keys: set[str] = {'username', 'password', 'excel.path', 'excel.sheet', 'download.path', 'rename.path',
                                    'excel.read_column.start_cell.fcr', 'excel.read_column.start_cell.fcr_rename'}
        return mandatory_keys

    def automate(self):
        logger: Logger = get_current_logger()
        logger.info(
            "---------------------------------------------------------------------------------------------------------")
        logger.info("Start processing")

        uid: str = self._settings['username']
        psw: str = self._settings['password']
        login_url = 'https://{}:{}@amerapps.apmoller.net/DutyDeduction/Grid.aspx?search=true'.format((uid), (psw))
        logger.info('Try to login')
        self._driver.get(login_url)
        logger.info("Login successfully")

        fcr_numbers: list[str] = get_excel_data_in_column_start_at_row(self._settings['excel.path'],
                                                                       self._settings['excel.sheet'],
                                                                       self._settings[
                                                                           'excel.read_column.start_cell.fcr'])

        needed_to_add_cookies = Duty.produce_needed_to_add_cookie_contents(batch_size=20, fcr_numbers=fcr_numbers)
        download_filter_cookies: list[str] = needed_to_add_cookies[0]
        search_filter_cookies: list[str] = needed_to_add_cookies[1]

        batch_index: int = 0
        quantity_of_batches: int = len(download_filter_cookies)
        while batch_index < quantity_of_batches:
            self._driver.add_cookie(
                {'name': 'Download1FilterString', 'value': f'{download_filter_cookies[batch_index]}'})
            self._driver.add_cookie(
                {'name': 'SearchControl1Download1Filter', 'value': f'{search_filter_cookies[batch_index]}'})
            self._driver.get(login_url)
            logger.info('Refreshed cookies')

            fcr_code_to_index_and_time: Dict[str, Tuple[int, datetime]] = {}

            fcr_elements: list[WebElement] = self._driver.find_elements(by=By.CSS_SELECTOR,
                                                                        value='table#EDIGrid.MyGrid tr td:nth-child(6) span')
            fcr_elements = fcr_elements[1:]

            fcr_index = 2
            for fcr_element in fcr_elements:

                fcr_code: str = fcr_element.get_attribute('innerText')
                date_element = self._driver.find_element(by=By.CSS_SELECTOR,
                                                         value=f'table#EDIGrid.MyGrid tr:nth-child({fcr_index}) '
                                                               f'td:nth-child(8) span')
                date_string = date_element.get_attribute('innerText')
                date_format = "%m/%d/%Y %I:%M:%S %p"
                fcr_datetime = datetime.strptime(date_string, date_format)

                if fcr_code not in fcr_code_to_index_and_time:
                    fcr_code_to_index_and_time[fcr_code] = (fcr_index, fcr_datetime)
                else:
                    last_fcr_datetime = fcr_code_to_index_and_time[fcr_code][1]
                    if fcr_datetime >= last_fcr_datetime:
                        fcr_code_to_index_and_time[fcr_code] = (fcr_index, fcr_datetime)

                fcr_index += 1

            for key, value in fcr_code_to_index_and_time.items():
                fcr_code = key
                fcr_index = value[0]
                self.click_download(fcr_code, fcr_index)
                self._rename_file_after_download(fcr_code, fcr_index)

            batch_index += 1

        logger.info("Complete download")
        self._input_excel()
        logger.info('Checked file exist - Check your file Excel to get infor')
        self._driver.close()
        logger.info(
            "---------------------------------------------------------------------------------------------------------")
        logger.info("End processing")

    def click_download(self, fcr_code: str, fcr_index: int):
        logger: Logger = get_current_logger()
        max_attempt = 5
        attempt_counter = 0
        while True:
            try:
                if attempt_counter >= max_attempt:
                    raise Exception('Can not click download for frc {}'.format(fcr_code))

                logger.info(f'Try to click on {fcr_code} at index {fcr_index}')
                self._click_when_element_present(by=By.CSS_SELECTOR,
                                                 value=f'table#EDIGrid.MyGrid tr:nth-child({fcr_index}) a')
                break
            except Exception:
                time.sleep(1)
                attempt_counter += 1

    @staticmethod
    def produce_needed_to_add_cookie_contents(batch_size: int = 20, fcr_numbers: list[str] = None) -> tuple[
        list[str], list[str]]:
        download_filter_cookies = []
        search_filter_cookies = []

        initial_download_filter_value = 'SavedString= and ('
        current_download_filter_cookie = initial_download_filter_value

        initial_search_filter_value = ''
        current_search_filter_cookie = initial_search_filter_value

        current_index = 0
        for fcr in fcr_numbers:
            current_search_filter_cookie += 'Download_Search_1_{}=SourceName&Download_Search_2_{}==&Download_Search_3_{}={}&'.format(
                current_index, current_index, current_index, fcr)
            current_download_filter_cookie += f'SourceName=\'{fcr}\' or '
            current_index += 1

            if current_index >= batch_size:
                current_download_filter_cookie = current_download_filter_cookie[:-4]
                download_filter_cookies.append(f'{current_download_filter_cookie})')
                current_download_filter_cookie = initial_download_filter_value

                current_search_filter_cookie = current_search_filter_cookie[:-1]
                search_filter_cookies.append(f'{current_search_filter_cookie})')
                current_search_filter_cookie = initial_search_filter_value

                current_index = 0

        if len(current_download_filter_cookie) > 0:
            current_download_filter_cookie = current_download_filter_cookie[:-4]
            download_filter_cookies.append(f'{current_download_filter_cookie})')

            current_search_filter_cookie = current_search_filter_cookie[:-1]
            search_filter_cookies.append(f'{current_search_filter_cookie})')

        return download_filter_cookies, search_filter_cookies

    def _rename_file_after_download(self, fcr_code: str, fcr_index: int):
        logger: Logger = get_current_logger()

        download_folder: str = self._settings['download.path']
        rename_folder: str = self._settings['rename.path']

        attempt_counter: int = 0
        max_attempt: int = 2 * 60
        while True:

            if attempt_counter > max_attempt:
                logger.error('The tool waiting too long to download document for {}. Please check !'.format(fcr_code))
                break

            all_files_in_download_folder: list[str] = os.listdir(download_folder)
            download_filename: str = None if len(all_files_in_download_folder) == 0 else all_files_in_download_folder[0]

            if download_filename is None or not download_filename.endswith('.pdf'):
                time.sleep(1)
                attempt_counter += 1
                continue

            # The download has completed
            logger.info('The document for {} has been downloaded !'.format(fcr_code))
            break

        # rename
        all_files_in_download_folder: list[str] = os.listdir(download_folder)
        download_filename: str = None if len(all_files_in_download_folder) == 0 else all_files_in_download_folder[0]

        rename_filename_path = os.path.join(rename_folder, '{}_Duty.pdf'.format(fcr_code))
        full_file_path: str = os.path.join(download_folder, download_filename)

        os.rename(full_file_path, rename_filename_path)
        logger.info('Renamed from {} to {}'.format(full_file_path, rename_filename_path))

    def _input_excel(self):
        logger: Logger = get_current_logger()

        wb = load_workbook(self._settings['excel.path'])
        ws = wb[self._settings['excel.sheet']]

        logger.info('Inputting Excel')
        rename_folder: str = self._settings['rename.path']

        new_status_done: str = 'Done'
        new_status_miss: str = 'Missing'

        all_file_in_rename_folders: list[str] = os.listdir(rename_folder)

        for index, cell in enumerate(ws['A'], start=1):
            cell_name = str(cell.value) + '_Duty.pdf'
            if cell_name in all_file_in_rename_folders:
                ws[f'B{index}'] = new_status_done
            else:
                ws[f'B{index}'] = new_status_miss

            new_file_path = os.path.join(rename_folder, 'output.xlsx')

            if os.path.exists(new_file_path):
                os.remove(new_file_path)
            wb.save(new_file_path)

    # _______________________________________________________________________________________________________________
    # def _rename_by_asp(self, postfix_url:str):
    #
    #     cookie_asp_session = self._driver.get_cookie('ASP.NET_SessionId')
    #     if cookie_asp_session:
    #         asp_session_value = cookie_asp_session.get('value')
    #
    #         if asp_session_value:
    #             pdf_url = postfix_url
    #             cookies = self._driver.get_cookies()
    #
    #             session = requests.Session()
    #
    #             cookies_string = ';'.join([f"{cookie['name']}={cookie['value']}" for cookie in cookies])
    #             headers = {'Cookie': f"NET_SessionId={asp_session_value}",
    #                        'Host' : 'amerapps.apmoller.net'}
    #
    #             session.auth = None
    #             response = session.get(pdf_url, headers=headers, verify=False, auth=None)
    #
    #             if response.status_code == 200:
    #                 # Lưu phản hồi từ yêu cầu requests dưới dạng tệp PDF
    #                 download_path = '/path/to/save/downloaded_file.pdf'  # Thay thế bằng đường dẫn lưu tệp PDF
    #                 with open(download_path, 'wb') as file:
    #                     file.write(response.content)
    #                 print(f"Download successful. File saved at: {download_path}")
    #             else:
    #                 print("Download failed.")
    #         else:
    #             print("ASPSession cookie value not found.")
    #     else:
    #         print("ASPSession cookie not found.")
    # ________________________________________________________________________________________________________________
